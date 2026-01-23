from django.contrib.gis.geos import Point
from openpyxl import load_workbook
from drf_spectacular.utils import extend_schema
from django.utils.timezone import make_aware
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status

from app.models import Event, Venue
from app.permissions import IsSuperUser
from app.serializers import EventImportSerializer


@extend_schema(
    request=EventImportSerializer,
    responses={201: None},
)
class EventImportView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [IsSuperUser]

    def post(self, request):
        serializer = EventImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        wb = load_workbook(serializer.validated_data["file"])
        ws = wb.active

        created = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            name, description, published_at, start_at, end_at, venue_name, lat, lon, rating = row

            venue, _ = Venue.objects.get_or_create(title=venue_name, defaults={"location": Point(lon, lat) })

            Event.objects.create(
                name=name,
                description=description,
                published_at=make_aware(published_at),
                start_at=make_aware(start_at),
                end_at=make_aware(end_at),
                venue=venue,
                rating=rating,
                author=request.user,
            )

            created += 1

        return Response({"created": created}, status=status.HTTP_201_CREATED)